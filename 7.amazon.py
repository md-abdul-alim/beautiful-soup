from bs4 import BeautifulSoup
import openpyxl

# Open the Amazon.html file and read its contents
with open('Amazon.html', "r", encoding='utf8') as f:
	html_content = f.read()

# parse the HTML content with Beatiful Soup
soup = BeautifulSoup(html_content, 'html.parser')

divs = soup.find_all("div", class_ = "s-card-container s-overflow-hidden aok-relative puis-include-content-margin puis puis-vqcnlkgcd4sr22kxlmcindh6i0 s-latency-cf-section s-card-border")
# print(divs)


# Open an Excel file and create a new worksheet

wb = openpyxl.Workbook()
ws = wb.active

# Write the column headers

ws['A1'] = "Product Name"
ws['B1'] = "Product Price"
ws['C1'] = "Product Reviews"

# Loopo through each div and extract the product name, price and reviews (if available)
row = 2

for div in divs:
	#Find the product name
	product_name_span = div.find("span", class_ = "a-size-medium a-color-base a-text-normal")
	if product_name_span:
		product_name = product_name_span.text.strip()
	else:
		product_name = "None"

	# Find the product price
	product_price_span = div.find("span", class_ = "a-price-whole")
	if product_price_span:
		product_price = product_price_span.text.strip()
	else:
		product_price = "None"

	# Find the product reviews
	product_reviews_span = div.find("span", class_ = "a-size-base s-underline-text")
	if product_reviews_span:
		product_review = product_reviews_span.text.strip()
	else:
		product_review = "None"

	# Write the data to the Excel worksheet
	ws.cell(row=row, column=1, value=product_name)
	ws.cell(row=row, column=2, value=product_price)
	ws.cell(row=row, column=3, value=product_review)

	row += 1

# Save the Excel file

wb.save("amazon_products.xlsx")